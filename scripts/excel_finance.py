def excel_finance_base():
    import openpyxl, re, csv
    from dotenv import dotenv_values

    # Read credentials from environment
    config = dotenv_values(".env")

    INPUT_FILE = config["INPUT_FILE"]
    LABEL_REP_FILE = config["LABEL_REP_FILE"]
    DESC_REP_FILE = config["DESC_REP_FILE"]
    WB_NAME = config["WB_NAME"]
    MONTH_END_GUARD = "[END]"
    CREDIT_ACCOUNT = "credit"
    UNCAT_LABEL = "[UNCAT]"
    DIV = "--------------------------------"
    def construct_dict_from_csv(INPUT_FILE):
        result = dict()
        with open(INPUT_FILE) as f:
            reader = csv.reader(f, delimiter = "\t")
            for row in reader:
                result[row[0]] = row[1]
        return result

    # Loads workbook
    workbook = openpyxl.Workbook()
    # Selects the first available sheet
    sheet = workbook.active

    # Save csv file as xlsx
    with open(INPUT_FILE) as f:
        reader = csv.reader(f)
        for row in reader:
            sheet.append(row)

    # Delete header
    HEADER = [c.value for c in sheet[1]]
    sheet.delete_rows(idx = 1, amount = 1)

    ROWS = sheet.max_row
    COLS = sheet.max_column

    # Filter for transactions after last month's end
    for n in range(ROWS): 
        if sheet.cell(row = n + 1, column = 2).value[:len(MONTH_END_GUARD)] == MONTH_END_GUARD:
            print("\nTruncating from last month end")
            print(f"{sheet.cell(row = n + 1, column = 1).value, sheet.cell(row = n + 1, column = 2).value}\n{DIV}\n")
            sheet.delete_rows(idx = n + 1, amount = ROWS - n)
            break
    ROWS = n

    # Convert all credit accounts to negative
    credit_count = 0
    for n in range(ROWS):
        amount = sheet.cell(row = n + 1, column = 4)
        if sheet.cell(row = n + 1, column = 5).value == CREDIT_ACCOUNT:
            amount.value = -float(amount.value)
            credit_count += 1
        else:
            amount.value = float(amount.value)
    print(f"Converted {credit_count} credit transactions to negative\n{DIV}\n")

    # Layer 1 Recategorization
    ## Construct label rep dict
    LABEL_REP = construct_dict_from_csv(LABEL_REP_FILE)

    ## Replace category labels
    recat_labels_count = 0
    uncat_labels_count = 0
    for n in range(ROWS):
        label = sheet.cell(row = n + 1, column = 6)
        if label.value in LABEL_REP:
            label.value = LABEL_REP[label.value]
            recat_labels_count += 1
        else:
            label.value = UNCAT_LABEL
            uncat_labels_count += 1

    print(f"Recategorized {recat_labels_count} transactions")
    print(f"Labeled {uncat_labels_count} uncategorizable transactions\n{DIV}\n")

    # Layer 2 Recategorization
    ## Construct desc rep dict
    DESC_REP = construct_dict_from_csv(DESC_REP_FILE)

    ## Replace category labels (with regex matching to description)
    desc_recat = 0
    for pattern, category in DESC_REP.items():
        for n in range(ROWS):
            desc = sheet.cell(row = n + 1, column = 3)
            label = sheet.cell(row = n + 1, column = 6)
            old_label = label.value
            if re.search(f'^({pattern}).*$', desc.value) and category != old_label:
                label.value = category
                desc_recat += 1
    print(f"Recategorized {desc_recat} transactions from description\n{DIV}\n")

    # Restore header
    sheet.insert_rows(1)
    for n in range(len(HEADER)):
        sheet.cell(row = 1, column =  n + 1).value = HEADER[n]
    ROWS = sheet.max_row

    # Remove original description header
    sheet.delete_cols(3)

    # Add aggregation sheet
    agg_sheet = workbook.create_sheet("Aggregation")

    ## Add formula to list unique categories, and sum if category matches
    agg_sheet["A1"] = f"=UNIQUE(Sheet!E2:E{ROWS})"
    for row in range(1,20):
        agg_sheet[f"B{row}"] = f"=SUMIF(Sheet!E:E, A{row}, Sheet!C:C)"
    print(f"Added aggregation\n{DIV}\n")

    print(f"Cleaning complete\n{DIV}\n")

    # Build final update template
    fin_sheet = workbook.create_sheet("Final")
    fin_header = ["CATEGORY", "AMOUNT", "ACCOUNT", "BALANCE"]
    for n in range(len(fin_header)):
        fin_sheet.cell(row = 1, column =  n + 1).value = fin_header[n]
    print(f"Added final template\n{DIV}\n")

    workbook.save(WB_NAME)