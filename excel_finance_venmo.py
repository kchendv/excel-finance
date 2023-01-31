import openpyxl, json, datetime, re, csv
from venmo_api import Client

# Read credentials from cred.json
with open('config.json') as f:
    config = json.load(f)

VENMO_AT = config["venmo_at"]
VENMO_PAY = "VENMO PAYMENT"
VENMO_RECIEVE = "VENMO CASHOUT"
SEND_TIME_THRESHOLD = 86400.0
REC_TIME_THRESHOLD = 345600.0
WB_NAME = "transactions_x.xlsx"
DIV = "--------------------------------"

# Venmo access instance
venmo = Client(access_token = VENMO_AT)

me_user = venmo.user.get_my_profile()

# Get / parse transactions
transactions = venmo.user.get_user_transactions(user = me_user) 

reduced_transactions = [(
    datetime.datetime.fromtimestamp(t.date_completed).strftime("%m/%#d/%Y"),
    t.actor.display_name,
    t.target.display_name,
    t.note,
    t.amount,
    datetime.datetime.fromtimestamp(t.date_completed))
    for t in transactions]

# Create venmo note category map
venmo_rep = dict()
with open("venmo_rep.txt") as f:
    reader = csv.reader(f, delimiter = "\t")
    for row in reader:
        venmo_rep[row[0]] = row[1]

# Attempt transaction matching / label replacement
wb = openpyxl.load_workbook(WB_NAME)
sheet = wb["Sheet"]

pay_count, recieve_count, pay_rep_count = 0, 0, 0
for n in range(sheet.max_row - 1): 
    # Payment message
    if re.search(f'^({VENMO_PAY}).*$', sheet.cell(row = n + 1, column = 2).value):
        t_date = datetime.datetime.strptime(sheet.cell(row = n + 1, column = 1).value, "%m/%d/%Y")
        t_amount = float(sheet.cell(row = n + 1, column = 3).value)
        for t in reduced_transactions:
            if abs((t_date - t[5]).total_seconds()) < SEND_TIME_THRESHOLD and t_amount == t[4]:
                # Replace message
                sheet.cell(row = n + 1, column = 2).value = f"VENMO PAY {t[2]} : [{t[3]}]"
                pay_count += 1

                # Attemp label matching
                for pattern, category in venmo_rep.items():
                    if re.search(f'^(.* )*({pattern}).*$', t[3]):
                        sheet.cell(row = n + 1, column = 4).value = category
                        pay_rep_count += 1
                        break
                break
    # Recieve message
    elif re.search(f'^({VENMO_RECIEVE}).*$', sheet.cell(row = n + 1, column = 2).value):
        t_date = datetime.datetime.strptime(sheet.cell(row = n + 1, column = 1).value, "%m/%d/%Y")
        t_amount = -float(sheet.cell(row = n + 1, column = 3).value)
        for t in reduced_transactions:
            if  abs((t_date - t[5]).total_seconds()) < REC_TIME_THRESHOLD and t_amount == t[4]:
                # Replace message
                sheet.cell(row = n + 1, column = 2).value = f"VENMO RECIEVE {t[1]} : [{t[3]}]"
                recieve_count += 1
                break

print(f"Updated description to {pay_count} pay / {recieve_count} recieve venmo transactions\n{DIV}\n")

print(f"Updated label to {pay_rep_count} venmo pay transactions\n{DIV}\n")

    # elif re.search(f'^({VENMO_RECIEVE}).*$', sheet.cell(row = n + 1, column = 2).value):
    #   t_date = sheet.cell(row = n + 1, column = 1).value.strptime("%m/%#d/%Y")
    #     for t in reduced_transactions:
    #         if abs(t_date - t[5]) < TIME_THRESHOLD:

# # Add header
# reduced_transactions = [("Date", "From", "To", "Note", "Amount")] + reduced_transactions

# # Fill in transaction data in workbook
# sheet = wb["Aggregation"]
# for i in range(len(reduced_transactions)):
#     for j in range(len(reduced_transactions[i]) - 1):
#         sheet.cell(row = i + 1, column = j + 4).value = reduced_transactions[i][j]
wb.save(WB_NAME)